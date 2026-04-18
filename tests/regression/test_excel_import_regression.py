from __future__ import annotations

import io

from openpyxl import Workbook

from server.modules.offers import map_excel_row_to_offer, preview_offers_from_excel, read_excel_rows_with_hyperlinks


def test_excel_import_does_not_return_formula_text() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "applications"
    worksheet.append(["Company", "Role"])
    worksheet["A2"] = "ACME"
    worksheet["B2"] = "=1+1"
    worksheet["B2"].hyperlink = "https://example.com/job"

    buf = io.BytesIO()
    workbook.save(buf)
    content = buf.getvalue()

    rows = read_excel_rows_with_hyperlinks(content)

    assert len(rows) == 1
    # For formula cells importer should not return raw '=...' expression.
    assert rows[0]["Role"] != "=1+1"
    assert rows[0]["__link__Role"] == "https://example.com/job"


def test_excel_import_uses_hyperlink_from_role_column_as_offer_url() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "applications"
    worksheet.append(["Firma", "Stanowisko", "Status", "Lokalizacja"])
    worksheet["A2"] = "ACME"
    worksheet["B2"] = "Python Developer"
    worksheet["B2"].hyperlink = "https://example.com/oferta/python-dev"
    worksheet["C2"] = "Wyslano"
    worksheet["D2"] = "Warszawa"

    buf = io.BytesIO()
    workbook.save(buf)
    content = buf.getvalue()

    rows = read_excel_rows_with_hyperlinks(content)
    mapped, issue = map_excel_row_to_offer(rows[0])

    assert issue is None
    assert mapped is not None
    assert mapped["role"] == "Python Developer"
    assert mapped["sourceUrl"] == "https://example.com/oferta/python-dev"


def test_excel_import_reads_separate_link_column_hyperlink() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "applications"
    worksheet.append(["Firma", "Stanowisko", "Status", "Lokalizacja", "Link"])
    worksheet["A2"] = "ACME"
    worksheet["B2"] = "Backend Developer"
    worksheet["C2"] = "Wyslano"
    worksheet["D2"] = "Warszawa"
    worksheet["E2"] = "Oferta"
    worksheet["E2"].hyperlink = "https://example.com/oferta/backend"

    buf = io.BytesIO()
    workbook.save(buf)
    content = buf.getvalue()

    rows = read_excel_rows_with_hyperlinks(content)
    mapped, issue = map_excel_row_to_offer(rows[0])

    assert issue is None
    assert mapped is not None
    assert mapped["role"] == "Backend Developer"
    assert mapped["sourceUrl"] == "https://example.com/oferta/backend"


def test_excel_import_reads_hyperlink_formula_from_role_cell() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "applications"
    worksheet.append(["Firma", "Stanowisko", "Status", "Lokalizacja"])
    worksheet["A2"] = "ACME"
    worksheet["B2"] = '=HYPERLINK("https://example.com/oferta/formula-role","Python Dev")'
    worksheet["C2"] = "Wyslano"
    worksheet["D2"] = "Warszawa"

    buf = io.BytesIO()
    workbook.save(buf)
    content = buf.getvalue()

    rows = read_excel_rows_with_hyperlinks(content)
    mapped, issue = map_excel_row_to_offer(rows[0])

    assert issue is None
    assert mapped is not None
    assert mapped["sourceUrl"] == "https://example.com/oferta/formula-role"


def test_excel_import_reads_localized_hyperlink_formula_in_company_cell() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "applications"
    worksheet.append(["Firma", "Stanowisko", "Status", "Lokalizacja"])
    worksheet["A2"] = '=HIPERLACZE("https://example.com/oferta/company-link";"Firma XYZ")'
    worksheet["B2"] = "Specjalista"
    worksheet["C2"] = "Wyslano"
    worksheet["D2"] = "Warszawa"

    buf = io.BytesIO()
    workbook.save(buf)
    content = buf.getvalue()

    rows = read_excel_rows_with_hyperlinks(content)
    mapped, issue = map_excel_row_to_offer(rows[0])

    assert issue is None
    assert mapped is not None
    assert mapped["company"] == "Firma XYZ"
    assert mapped["sourceUrl"] == "https://example.com/oferta/company-link"


def test_excel_import_reads_company_name_header_variant_and_hyperlink_formula() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "applications"
    worksheet.append(["Company Name", "Nazwa stanowiska", "Status", "Lokalizacja"])
    worksheet["A2"] = '=IFERROR(HYPERLINK("https://example.com/oferta/company-variant","Firma Variant"),"")'
    worksheet["B2"] = "QA Engineer"
    worksheet["C2"] = "Wyslano"
    worksheet["D2"] = "Łódź"

    buf = io.BytesIO()
    workbook.save(buf)
    content = buf.getvalue()

    rows = read_excel_rows_with_hyperlinks(content)
    mapped, issue = map_excel_row_to_offer(rows[0])

    assert issue is None
    assert mapped is not None
    assert mapped["company"] == "Firma Variant"
    assert mapped["role"] == "QA Engineer"
    assert mapped["sourceUrl"] == "https://example.com/oferta/company-variant"


def test_excel_import_reads_company_from_merged_cell() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "applications"
    worksheet.append(["Firma", "Stanowisko", "Status", "Lokalizacja"])
    worksheet["A2"] = "Epufloor Sp. z o.o."
    worksheet.merge_cells("A2:A3")
    worksheet["B2"] = "Pracownik biurowy"
    worksheet["C2"] = "Wyslano"
    worksheet["D2"] = "Łódź"
    worksheet["B3"] = "Pracownik biurowy w dziale logistyki (k/m)"
    worksheet["C3"] = "Wyslano"
    worksheet["D3"] = "Łódź"

    buf = io.BytesIO()
    workbook.save(buf)
    content = buf.getvalue()

    preview = preview_offers_from_excel(content)

    assert preview["skipped"] == 0
    assert preview["ignored"] == 0
    assert preview["imported"] == 2
    assert preview["offers"][1]["company"] == "Epufloor Sp. z o.o."
    assert preview["offers"][1]["role"] == "Pracownik biurowy w dziale logistyki (k/m)"


def test_excel_import_does_not_copy_company_when_cell_is_plain_empty() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "applications"
    worksheet.append(["Firma", "Stanowisko", "Status", "Lokalizacja"])
    worksheet["A2"] = "Epufloor Sp. z o.o."
    worksheet["B2"] = "Pracownik biurowy"
    worksheet["C2"] = "Wyslano"
    worksheet["D2"] = "Łódź"
    worksheet["A3"] = None
    worksheet["B3"] = "Pracownik biurowy w dziale logistyki (k/m)"
    worksheet["C3"] = "Wyslano"
    worksheet["D3"] = "Łódź"

    buf = io.BytesIO()
    workbook.save(buf)
    content = buf.getvalue()

    preview = preview_offers_from_excel(content)

    assert preview["imported"] == 1
    assert preview["skipped"] == 1
    assert preview["offers"][0]["company"] == "Epufloor Sp. z o.o."


def test_excel_import_resolves_simple_company_concat_formula() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "applications"
    worksheet.append(["company", "role", "status", "location"])
    worksheet["A2"] = "Epufloor Sp. z o.o."
    worksheet["A3"] = '="Przedsiebiorstwo " & A2'
    worksheet["B2"] = "Pracownik biurowy"
    worksheet["B3"] = "Pracownik biurowy w dziale logistyki (k/m)"
    worksheet["C2"] = "Wyslano"
    worksheet["C3"] = "Wyslano"
    worksheet["D2"] = "Lodz"
    worksheet["D3"] = "Lodz"

    buf = io.BytesIO()
    workbook.save(buf)
    content = buf.getvalue()

    preview = preview_offers_from_excel(content)

    assert preview["imported"] == 2
    assert preview["skipped"] == 0
    assert preview["offers"][1]["company"] == "Przedsiebiorstwo Epufloor Sp. z o.o."


def test_excel_import_ignores_circular_company_formula_without_fake_value() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "applications"
    worksheet.append(["company", "role", "status", "location"])
    worksheet["A2"] = '="Przedsiebiorstwo " & A3'
    worksheet["A3"] = '="Przedsiebiorstwo " & A2'
    worksheet["B2"] = "Rola A"
    worksheet["B3"] = "Rola B"
    worksheet["C2"] = "Wyslano"
    worksheet["C3"] = "Wyslano"
    worksheet["D2"] = "Lodz"
    worksheet["D3"] = "Lodz"

    buf = io.BytesIO()
    workbook.save(buf)
    content = buf.getvalue()

    preview = preview_offers_from_excel(content)

    assert preview["imported"] == 0
    assert preview["skipped"] == 2
    assert len(preview["issues"]) == 2
