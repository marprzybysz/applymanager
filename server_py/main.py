from __future__ import annotations

import io
import os
import re
from datetime import date, datetime, timezone
from typing import Any
from urllib.parse import urlparse

import psycopg2
from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel
from psycopg2.extras import RealDictCursor

from server_py.scrapers.index import get_supported_sources, scrape_job_from_link, scrape_jobs

PORT = int(os.getenv("PORT", "3000"))
DIST_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "dist"))

DB_CONFIG = {
    "host": os.getenv("DB_HOST", "localhost"),
    "port": int(os.getenv("DB_PORT", "5432")),
    "user": os.getenv("DB_USER", "applymanager"),
    "password": os.getenv("DB_PASSWORD", "applymanager"),
    "dbname": os.getenv("DB_NAME", "applymanager"),
}

EXCEL_FIELDS = {
    "company": ["company", "Company", "firma", "Firma"],
    "role": ["role", "Role", "position", "Position", "stanowisko", "Stanowisko"],
    "applied": ["applied", "Applied", "aplikowano", "Aplikowano"],
    "status": ["status", "Status"],
    "location": ["location", "Location", "lokalizacja", "Lokalizacja"],
    "notes": ["notes", "Notes", "notatki", "Notatki"],
    "appliedAt": ["applied_at", "appliedAt", "date", "Date", "data", "Data", "data aplikacji", "Data aplikacji"],
    "source": ["source", "Source", "portal", "Portal"],
    "sourceUrl": ["url", "URL", "link", "Link", "hyperlink", "Hyperlink"],
    "sourceUrlLink": ["__link__url", "__link__URL", "__link__link", "__link__Link", "__link__hyperlink", "__link__Hyperlink"],
}

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


def get_connection():
    return psycopg2.connect(**DB_CONFIG)


def to_non_empty_string(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def to_boolean_or_null(value: Any) -> bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        if int(value) == 1:
            return True
        if int(value) == 0:
            return False
        return None

    normalized = str(value).strip().lower()
    if not normalized:
        return None
    if normalized in {"true", "1", "yes", "y", "tak", "t", "applied"}:
        return True
    if normalized in {"false", "0", "no", "n", "nie", "f", "not_applied", "not-applied"}:
        return False
    return None


def safe_date(value: Any) -> str | None:
    if value in (None, ""):
        return None

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, (int, float)):
        try:
            parsed = from_excel(value)
            if isinstance(parsed, datetime):
                return parsed.date().isoformat()
            if isinstance(parsed, date):
                return parsed.isoformat()
        except Exception:
            pass

    text = str(value).strip()
    if not text:
        return None

    if text.isdigit():
        try:
            parsed = from_excel(float(text))
            if isinstance(parsed, datetime):
                return parsed.date().isoformat()
            if isinstance(parsed, date):
                return parsed.isoformat()
        except Exception:
            pass

    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date().isoformat()
    except Exception:
        return None


def is_absolute_http_url(value: Any) -> bool:
    try:
        parsed = urlparse(str(value))
        return parsed.scheme in {"http", "https"} and bool(parsed.netloc)
    except Exception:
        return False


def pick_first_value(row: dict[str, Any], candidates: list[str]) -> str | None:
    for key in candidates:
        value = to_non_empty_string(row.get(key))
        if value:
            return value
    return None


def map_excel_row_to_offer(row: dict[str, Any]) -> dict[str, Any] | None:
    company = pick_first_value(row, EXCEL_FIELDS["company"])
    role = pick_first_value(row, EXCEL_FIELDS["role"])
    if not company or not role:
        return None

    direct_source_url = pick_first_value(row, EXCEL_FIELDS["sourceUrl"])
    linked_source_url = pick_first_value(row, EXCEL_FIELDS["sourceUrlLink"])
    applied = to_boolean_or_null(pick_first_value(row, EXCEL_FIELDS["applied"]))
    applied_value = True if applied is None else applied
    explicit_status = pick_first_value(row, EXCEL_FIELDS["status"])

    return {
        "company": company,
        "role": role,
        "applied": applied_value,
        "status": explicit_status or ("applied" if applied_value else "saved"),
        "location": pick_first_value(row, EXCEL_FIELDS["location"]),
        "notes": pick_first_value(row, EXCEL_FIELDS["notes"]),
        "appliedAt": safe_date(pick_first_value(row, EXCEL_FIELDS["appliedAt"])),
        "source": pick_first_value(row, EXCEL_FIELDS["source"]),
        "sourceUrl": direct_source_url if is_absolute_http_url(direct_source_url) else (linked_source_url or None),
    }


def read_excel_rows_with_hyperlinks(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(content), data_only=False)
    worksheet = workbook.worksheets[0]

    header_cells = list(next(worksheet.iter_rows(min_row=1, max_row=1), []))
    headers = [to_non_empty_string(cell.value) or "" for cell in header_cells]
    rows: list[dict[str, Any]] = []

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        item: dict[str, Any] = {}
        for idx, cell in enumerate(row):
            if idx >= len(headers):
                continue
            header = headers[idx]
            if not header:
                continue
            item[header] = cell.value
            if cell.hyperlink and cell.hyperlink.target:
                item[f"__link__{header}"] = str(cell.hyperlink.target).strip()
        rows.append(item)

    return rows


def extract_first_http_url(input_text: Any) -> str | None:
    text = to_non_empty_string(input_text)
    if not text:
        return None

    match = re.search(r"https?://[^\s\"'<>]+", text, flags=re.IGNORECASE)
    if not match:
        return None

    return re.sub(r"[),.;!?]+$", "", match.group(0))


def normalize_scrape_url(input_text: Any) -> str | None:
    candidate = extract_first_http_url(input_text) or to_non_empty_string(input_text)
    if not candidate or not is_absolute_http_url(candidate):
        return None

    parsed = urlparse(candidate)
    host = (parsed.hostname or "").lower()

    if "pracuj.pl" in host:
        return f"{parsed.scheme}://{parsed.netloc}{parsed.path}"
    return candidate


def map_offer_for_insert_from_request(body: dict[str, Any]) -> dict[str, Any]:
    applied = to_boolean_or_null(body.get("applied"))
    applied_value = True if applied is None else applied

    return {
        "company": to_non_empty_string(body.get("company")) or "",
        "role": to_non_empty_string(body.get("role")) or "",
        "applied": applied_value,
        "status": to_non_empty_string(body.get("status")) or ("applied" if applied_value else "saved"),
        "location": to_non_empty_string(body.get("location")),
        "notes": to_non_empty_string(body.get("notes")),
        "appliedAt": body.get("appliedAt"),
        "source": to_non_empty_string(body.get("source")),
        "sourceUrl": to_non_empty_string(body.get("sourceUrl")),
    }


def ensure_schema() -> None:
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS applications (
                  id SERIAL PRIMARY KEY,
                  company TEXT NOT NULL,
                  role TEXT NOT NULL,
                  applied BOOLEAN NOT NULL DEFAULT TRUE,
                  status TEXT NOT NULL DEFAULT 'applied',
                  location TEXT,
                  notes TEXT,
                  applied_at DATE,
                  source TEXT,
                  source_url TEXT,
                  created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                  updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                )
                """
            )
            cur.execute("ALTER TABLE applications ADD COLUMN IF NOT EXISTS applied BOOLEAN")
            cur.execute("UPDATE applications SET applied = TRUE WHERE applied IS NULL")
            cur.execute("ALTER TABLE applications ALTER COLUMN applied SET DEFAULT TRUE")
            cur.execute("ALTER TABLE applications ALTER COLUMN applied SET NOT NULL")
            cur.execute("ALTER TABLE applications ADD COLUMN IF NOT EXISTS source TEXT")
            cur.execute("ALTER TABLE applications ADD COLUMN IF NOT EXISTS source_url TEXT")
        conn.commit()


def list_offers() -> list[dict[str, Any]]:
    with get_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT
                    id,
                    company,
                    role,
                    applied,
                    status,
                    location,
                    notes,
                    applied_at AS "appliedAt",
                    source,
                    source_url AS "sourceUrl",
                    created_at AS "createdAt"
                FROM applications
                ORDER BY created_at DESC
                LIMIT 500
                """
            )
            rows = cur.fetchall()
            return [dict(row) for row in rows]


def insert_offer(offer: dict[str, Any]) -> dict[str, Any]:
    with get_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                INSERT INTO applications (company, role, applied, status, location, notes, applied_at, source, source_url)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id, company, role, applied, status, location, notes,
                          applied_at AS "appliedAt", source, source_url AS "sourceUrl", created_at AS "createdAt"
                """,
                (
                    offer["company"],
                    offer["role"],
                    offer.get("applied", True),
                    offer.get("status") or ("saved" if offer.get("applied") is False else "applied"),
                    offer.get("location") or None,
                    offer.get("notes") or None,
                    safe_date(offer.get("appliedAt")),
                    offer.get("source") or None,
                    offer.get("sourceUrl") or None,
                ),
            )
            row = cur.fetchone()
        conn.commit()
        return dict(row) if row else {}


@app.on_event("startup")
def startup_event() -> None:
    ensure_schema()


@app.get("/api/greet")
def greet(name: str | None = None):
    username = name.strip() if isinstance(name, str) and name.strip() else "friend"
    return {"message": f"Hello, {username}! Python FastAPI backend is connected."}


@app.get("/api/health")
def health():
    try:
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT NOW()")
                now = cur.fetchone()[0]
        return {"ok": True, "db": "connected", "now": now}
    except Exception as error:
        return JSONResponse(status_code=500, content={"ok": False, "db": "disconnected", "error": str(error)})


@app.get("/api/offers")
def get_offers():
    try:
        return {"ok": True, "offers": list_offers()}
    except Exception as error:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(error)})


@app.post("/api/offers")
async def create_offer(request: Request):
    body = await request.json()
    offer_input = map_offer_for_insert_from_request(body or {})

    if not offer_input["company"] or not offer_input["role"]:
        raise HTTPException(status_code=400, detail="company and role are required")

    try:
        offer = insert_offer(offer_input)
        return JSONResponse(status_code=201, content={"ok": True, "offer": offer})
    except Exception as error:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(error)})


@app.post("/api/offers/import-excel")
async def import_excel(file: UploadFile = File(...)):
    try:
        content = await file.read()
        rows = read_excel_rows_with_hyperlinks(content)
        mapped_offers = [offer for offer in (map_excel_row_to_offer(row) for row in rows) if offer]

        saved: list[dict[str, Any]] = []
        for offer in mapped_offers:
            saved.append(insert_offer(offer))

        return {
            "ok": True,
            "imported": len(saved),
            "skipped": len(rows) - len(saved),
            "offers": saved,
        }
    except Exception as error:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(error)})


@app.get("/api/offers/export-excel")
def export_excel():
    try:
        offers = list_offers()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "applications"

        headers = ["company", "role", "applied", "status", "location", "notes", "appliedAt", "source", "sourceUrl", "createdAt"]
        worksheet.append(headers)

        for offer in offers:
            worksheet.append(
                [
                    offer.get("company") or "",
                    offer.get("role") or "",
                    offer.get("applied") if offer.get("applied") is not None else True,
                    offer.get("status") or "",
                    offer.get("location") or "",
                    offer.get("notes") or "",
                    safe_date(offer.get("appliedAt")) or "",
                    offer.get("source") or "",
                    offer.get("sourceUrl") or "",
                    str(offer.get("createdAt") or ""),
                ]
            )

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        filename = f"applymanager-offers-{datetime.now(timezone.utc).date().isoformat()}.xlsx"
        headers = {
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
        return Response(
            content=buffer.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
    except Exception as error:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(error)})


@app.get("/api/scrape/sources")
def scrape_sources():
    return {"sources": get_supported_sources()}


@app.post("/api/scrape")
async def scrape(request: Request):
    body = await request.json()
    query = to_non_empty_string((body or {}).get("query")) or ""
    sources = (body or {}).get("sources") if isinstance((body or {}).get("sources"), list) else None
    limit_per_source = (body or {}).get("limitPerSource", 20)

    if not query:
        return JSONResponse(status_code=400, content={"ok": False, "error": "query is required"})

    try:
        normalized_query_url = normalize_scrape_url(query)
        if normalized_query_url:
            job = scrape_job_from_link(normalized_query_url)
            return {
                "ok": True,
                "mode": "link",
                "query": normalized_query_url,
                "total": 1,
                "sources": [{"source": job.get("source"), "ok": True, "jobs": [job], "fetchedFrom": normalized_query_url, "count": 1}],
                "jobs": [job],
            }

        result = scrape_jobs(query=query, sources=sources, limit_per_source=limit_per_source)
        return {"ok": True, "mode": "search", **result}
    except Exception as error:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(error)})


@app.post("/api/scrape/link")
async def scrape_link(request: Request):
    body = await request.json()
    url = normalize_scrape_url((body or {}).get("url")) or ""

    if not url:
        return JSONResponse(status_code=400, content={"ok": False, "error": "url is required"})

    try:
        parsed = scrape_job_from_link(url)
        return {"ok": True, "job": parsed}
    except Exception as error:
        return JSONResponse(status_code=400, content={"ok": False, "error": str(error)})


if os.path.isdir(DIST_DIR):
    app.mount("/", StaticFiles(directory=DIST_DIR, html=True), name="dist")


@app.get("/{full_path:path}")
def spa_fallback(full_path: str):
    index_file = os.path.join(DIST_DIR, "index.html")
    if os.path.exists(index_file):
        return FileResponse(index_file)
    raise HTTPException(status_code=404, detail=f"Not found: {full_path}")
