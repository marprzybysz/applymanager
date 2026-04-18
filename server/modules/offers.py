from __future__ import annotations

import io
import re
from datetime import date, datetime, timedelta, timezone
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from psycopg2.extras import RealDictCursor

from server.modules.common import is_absolute_http_url, safe_date, to_boolean_or_null, to_non_empty_string
from server.modules.db import get_connection

EXCEL_FIELDS = {
    "company": ["company", "Company", "company name", "Company name", "companyName", "firma", "Firma", "nazwa firmy", "Nazwa firmy", "pracodawca", "Pracodawca"],
    "role": ["role", "Role", "position", "Position", "job title", "Job title", "stanowisko", "Stanowisko", "nazwa stanowiska", "Nazwa stanowiska"],
    "applied": ["applied", "Applied", "aplikowano", "Aplikowano"],
    "status": ["status", "Status"],
    "location": ["location", "Location", "lokalizacja", "Lokalizacja"],
    "notes": ["notes", "Notes", "notatki", "Notatki"],
    "appliedAt": ["applied_at", "appliedAt", "date", "Date", "data", "Data", "data aplikacji", "Data aplikacji"],
    "datePosted": [
        "date_posted",
        "datePosted",
        "postedAt",
        "publishedAt",
        "date posted",
        "Date posted",
        "data publikacji",
        "Data publikacji",
        "opublikowano",
        "Opublikowano",
    ],
    "expiresAt": [
        "expires_at",
        "expiresAt",
        "validThrough",
        "valid through",
        "expiryDate",
        "expirationDate",
        "closingDate",
        "deadline",
        "wygasa",
        "Wygasa",
        "wygasa dnia",
        "Wygasa dnia",
        "data wygasniecia",
        "Data wygasniecia",
        "data wygaśnięcia",
        "Data wygaśnięcia",
        "data zamkniecia",
        "Data zamkniecia",
        "data zamknięcia",
        "Data zamknięcia",
        "termin",
        "Termin",
    ],
    "source": ["source", "Source", "portal", "Portal"],
    "archive": ["archive", "Archive", "archiwum", "Archiwum", "archived", "Archived", "zarchiwizowano", "Zarchiwizowano"],
    "sourceUrl": ["url", "URL", "link", "Link", "hyperlink", "Hyperlink", "link oferty", "Link oferty"],
    "sourceUrlLink": [
        "__link__url",
        "__link__URL",
        "__link__link",
        "__link__Link",
        "__link__hyperlink",
        "__link__Hyperlink",
        "__link__link oferty",
        "__link__Link oferty",
    ],
}
DEFAULT_OFFER_DURATION_DAYS = 30


def _normalize_header_name(value: Any) -> str:
    raw = to_non_empty_string(value)
    if not raw:
        return ""
    base = raw.strip().lower()
    replacements = str.maketrans(
        {
            "ą": "a",
            "ć": "c",
            "ę": "e",
            "ł": "l",
            "ń": "n",
            "ó": "o",
            "ś": "s",
            "ź": "z",
            "ż": "z",
        }
    )
    base = base.translate(replacements)
    # Normalize separators and punctuation so "Company Name", "company_name" and "company-name" match.
    base = re.sub(r"[^a-z0-9]+", " ", base).strip()
    return re.sub(r"\s+", " ", base)


def _find_matching_row_key(row: dict[str, Any], candidates: list[str]) -> str | None:
    for key in candidates:
        if key in row:
            return key

    normalized_to_original: dict[str, str] = {}
    for row_key in row.keys():
        normalized = _normalize_header_name(row_key)
        if normalized and normalized not in normalized_to_original:
            normalized_to_original[normalized] = str(row_key)

    for key in candidates:
        normalized = _normalize_header_name(key)
        if normalized and normalized in normalized_to_original:
            return normalized_to_original[normalized]
    return None


def pick_first_value_with_key(row: dict[str, Any], candidates: list[str]) -> tuple[str | None, str | None]:
    key = _find_matching_row_key(row, candidates)
    if not key:
        return None, None
    value = to_non_empty_string(row.get(key))
    return value, key


def pick_first_value(row: dict[str, Any], candidates: list[str]) -> str | None:
    value, _ = pick_first_value_with_key(row, candidates)
    return value


def _json_safe_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def _is_identifier_header(header: str) -> bool:
    normalized = str(header or "").strip().lower()
    return normalized in {"id", "lp", "nr", "no", "index", "row", "__rownumber"}


def _has_meaningful_non_id_data(row: dict[str, Any]) -> bool:
    for key, value in row.items():
        key_text = str(key or "")
        if key_text.startswith("__link__"):
            continue
        if _is_identifier_header(key_text):
            continue
        if to_non_empty_string(value):
            return True
    return False


def _normalize_status_key(value: Any) -> str:
    return str(value or "").strip().lower()


def _extract_hyperlink_formula_parts(value: Any) -> tuple[str | None, str | None]:
    text = to_non_empty_string(value)
    if not text:
        return None, None
    # Handles formulas like:
    # =HYPERLINK("https://example.com","Role")
    # =IFERROR(HYPERLINK("https://example.com";"Role"),"")
    # and locale variants HIPERLACZE / HIPERŁĄCZE.
    match = re.search(
        r'(?:_xlfn\.)?(?:HYPERLINK|HIPERLACZE|HIPERŁĄCZE)\s*\(\s*"([^"]+)"(?:\s*[,;]\s*"([^"]*)")?',
        text,
        flags=re.IGNORECASE,
    )
    if not match:
        return None, None
    url = to_non_empty_string(match.group(1))
    label = to_non_empty_string(match.group(2))
    if not (url and is_absolute_http_url(url)):
        return None, label
    return url, label


def _split_excel_concat_parts(expression: str) -> list[str]:
    parts: list[str] = []
    current: list[str] = []
    in_quotes = False
    i = 0
    while i < len(expression):
        ch = expression[i]
        if ch == '"':
            current.append(ch)
            if in_quotes and i + 1 < len(expression) and expression[i + 1] == '"':
                # Escaped quote in Excel string literal.
                current.append('"')
                i += 2
                continue
            in_quotes = not in_quotes
            i += 1
            continue
        if ch == "&" and not in_quotes:
            part = "".join(current).strip()
            if part:
                parts.append(part)
            current = []
            i += 1
            continue
        current.append(ch)
        i += 1

    tail = "".join(current).strip()
    if tail:
        parts.append(tail)
    return parts


def _normalize_cell_ref(token: str) -> str | None:
    match = re.match(r"^\$?([A-Z]{1,3})\$?(\d+)$", token.strip(), flags=re.IGNORECASE)
    if not match:
        return None
    return f"{match.group(1).upper()}{match.group(2)}"


def _evaluate_text_formula_cell(
    worksheet_values: Any,
    worksheet_links: Any,
    cell_ref: str,
    visiting: set[str],
) -> str | None:
    normalized_ref = _normalize_cell_ref(cell_ref)
    if not normalized_ref:
        return None
    if normalized_ref in visiting:
        # Circular dependency.
        return None

    visiting.add(normalized_ref)
    try:
        value_cell = worksheet_values[normalized_ref]
        if value_cell.value is not None:
            return str(value_cell.value)

        formula_cell = worksheet_links[normalized_ref]
        formula_text = to_non_empty_string(formula_cell.value)
        if not formula_text or not formula_text.startswith("="):
            return None

        expression = formula_text[1:].strip()
        parts = _split_excel_concat_parts(expression)
        if not parts:
            return None

        resolved: list[str] = []
        for part in parts:
            if re.match(r'^".*"$', part, flags=re.DOTALL):
                # Strip outer quotes and unescape doubled quotes.
                literal = part[1:-1].replace('""', '"')
                resolved.append(literal)
                continue

            reference = _normalize_cell_ref(part)
            if not reference:
                # Unsupported formula token (function, arithmetic, etc.).
                return None

            nested = _evaluate_text_formula_cell(worksheet_values, worksheet_links, reference, visiting)
            if nested is None:
                return None
            resolved.append(nested)

        return "".join(resolved)
    finally:
        visiting.remove(normalized_ref)


def _pick_role_hyperlink(row: dict[str, Any], role_key: str | None = None) -> str | None:
    if role_key:
        link_key = f"__link__{role_key}"
        link_value = to_non_empty_string(row.get(link_key))
        if link_value and is_absolute_http_url(link_value):
            return link_value

    for role_header in EXCEL_FIELDS["role"]:
        link_key = f"__link__{role_header}"
        link_value = to_non_empty_string(row.get(link_key))
        if link_value and is_absolute_http_url(link_value):
            return link_value
    return None


def _pick_any_row_hyperlink(row: dict[str, Any]) -> str | None:
    for key, value in row.items():
        key_text = str(key or "")
        if not key_text.startswith("__link__"):
            continue
        link_value = to_non_empty_string(value)
        if link_value and is_absolute_http_url(link_value):
            return link_value
    return None


def normalize_offer_status(status_value: Any, applied_default: bool = True) -> str:
    normalized = _normalize_status_key(status_value)
    if not normalized:
        return "Wyslano" if applied_default else "Zapisano"

    if normalized in {"applied", "wyslano", "wysłano", "sent", "zaaplikowano"}:
        return "Wyslano"
    if normalized in {"saved", "zapisano", "draft"}:
        return "Zapisano"
    if normalized in {"odczytano", "odczytana", "read"}:
        return "Odczytano"
    if normalized in {"interview", "in progress", "w trakcie", "proces"}:
        return "W trakcie"
    if normalized in {"rozmowa", "rozmowa umowiona", "umowienie na rozmowe", "umówienie na rozmowę"}:
        return "Rozmowa"
    if normalized in {"offer", "oferta"}:
        return "Oferta"
    if "rejected" in normalized or "odrzu" in normalized:
        return "Odrzucono"
    if "odmow" in normalized:
        return "Odmowa"

    return "Wyslano" if applied_default else "Zapisano"


def infer_date_posted_from_expires_at(expires_at: str | None) -> str | None:
    if not expires_at:
        return None
    try:
        inferred = date.fromisoformat(expires_at) - timedelta(days=DEFAULT_OFFER_DURATION_DAYS)
        return inferred.isoformat()
    except Exception:
        return None


def _format_created_at_for_export(value: Any, user_utc_offset_minutes: int | None) -> str:
    if value is None:
        return ""

    if isinstance(value, datetime):
        dt = value
    else:
        text = str(value).strip()
        if not text:
            return ""
        try:
            dt = datetime.fromisoformat(text.replace("Z", "+00:00"))
        except Exception:
            return text

    if user_utc_offset_minutes is None:
        return dt.isoformat()

    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)

    target_tz = timezone(timedelta(minutes=user_utc_offset_minutes))
    return dt.astimezone(target_tz).isoformat(timespec="seconds")


def map_offer_for_insert_from_request(body: dict[str, Any]) -> dict[str, Any]:
    applied = to_boolean_or_null(body.get("applied"))
    applied_value = True if applied is None else applied
    archived = to_boolean_or_null(body.get("archive"))
    archive_value = False if archived is None else archived

    return {
        "company": to_non_empty_string(body.get("company")) or "",
        "role": to_non_empty_string(body.get("role")) or "",
        "applied": applied_value,
        "archive": archive_value,
        "status": normalize_offer_status(body.get("status"), applied_value),
        "location": to_non_empty_string(body.get("location")),
        "notes": to_non_empty_string(body.get("notes")),
        "appliedAt": body.get("appliedAt"),
        "datePosted": body.get("datePosted"),
        "expiresAt": body.get("expiresAt"),
        "source": to_non_empty_string(body.get("source")),
        "sourceUrl": to_non_empty_string(body.get("sourceUrl")),
        "employmentTypes": [str(v).strip() for v in (body.get("employmentTypes") or []) if str(v).strip()],
        "workTime": to_non_empty_string(body.get("workTime")),
        "workMode": to_non_empty_string(body.get("workMode")),
        "shiftCount": to_non_empty_string(body.get("shiftCount")),
        "workingHours": to_non_empty_string(body.get("workingHours")),
    }


def map_excel_row_to_offer(
    row: dict[str, Any], import_source: str = "import_excel"
) -> tuple[dict[str, Any] | None, dict[str, Any] | None]:
    # Ignore empty rows or rows that only contain an ID-like value.
    if not _has_meaningful_non_id_data(row):
        return None, None

    company, company_key = pick_first_value_with_key(row, EXCEL_FIELDS["company"])
    role, role_key = pick_first_value_with_key(row, EXCEL_FIELDS["role"])

    role_hyperlink = _pick_role_hyperlink(row, role_key=role_key)
    company_hyperlink = _pick_role_hyperlink(row, role_key=company_key)
    direct_source_url = pick_first_value(row, EXCEL_FIELDS["sourceUrl"])
    linked_source_url = pick_first_value(row, EXCEL_FIELDS["sourceUrlLink"])
    source_url = (
        role_hyperlink
        or company_hyperlink
        or linked_source_url
        or (direct_source_url if is_absolute_http_url(direct_source_url) else None)
        or _pick_any_row_hyperlink(row)
    )
    applied = to_boolean_or_null(pick_first_value(row, EXCEL_FIELDS["applied"]))
    applied_value = True if applied is None else applied
    explicit_status = pick_first_value(row, EXCEL_FIELDS["status"])

    mapped_offer = {
        "company": company,
        "role": role,
        "applied": applied_value,
        "archive": to_boolean_or_null(pick_first_value(row, EXCEL_FIELDS["archive"])) is True,
        "status": normalize_offer_status(explicit_status, applied_value),
        "location": pick_first_value(row, EXCEL_FIELDS["location"]),
        "notes": pick_first_value(row, EXCEL_FIELDS["notes"]),
        "appliedAt": safe_date(pick_first_value(row, EXCEL_FIELDS["appliedAt"])),
        "datePosted": safe_date(pick_first_value(row, EXCEL_FIELDS["datePosted"])),
        "expiresAt": safe_date(pick_first_value(row, EXCEL_FIELDS["expiresAt"])),
        "source": pick_first_value(row, EXCEL_FIELDS["source"]) or import_source,
        "sourceUrl": source_url,
    }

    missing_fields: list[str] = []
    formula_issues: list[str] = []
    if not company:
        missing_fields.append("company")
        if company_key:
            formula_value = to_non_empty_string(row.get(f"__formula__{company_key}"))
            if formula_value and formula_value.startswith("="):
                formula_issues.append("unresolved_company_formula")
    if not role:
        missing_fields.append("role")

    issue = None
    if missing_fields:
        issue = {
            "rowNumber": row.get("__rowNumber"),
            "missingFields": missing_fields,
            "parsed": {
                "company": mapped_offer.get("company"),
                "role": mapped_offer.get("role"),
                "status": mapped_offer.get("status"),
                "location": mapped_offer.get("location"),
                "datePosted": mapped_offer.get("datePosted"),
                "expiresAt": mapped_offer.get("expiresAt"),
                "source": mapped_offer.get("source"),
                "sourceUrl": mapped_offer.get("sourceUrl"),
                "notes": mapped_offer.get("notes"),
            },
            "raw": {k: _json_safe_value(v) for k, v in row.items() if not str(k).startswith("__link__")},
        }
        if formula_issues:
            issue["formulaIssues"] = formula_issues

    if missing_fields:
        return None, issue
    return mapped_offer, issue


def read_excel_rows_with_hyperlinks(content: bytes) -> list[dict[str, Any]]:
    # Read values with data_only=True so formula cells return computed value (if cached in file),
    # while reading hyperlinks from a second workbook loaded with formulas preserved.
    workbook_values = load_workbook(io.BytesIO(content), data_only=True)
    workbook_links = load_workbook(io.BytesIO(content), data_only=False)
    worksheet_values = workbook_values.worksheets[0]
    worksheet_links = workbook_links.worksheets[0]

    merged_anchor_by_position: dict[tuple[int, int], tuple[int, int]] = {}
    for merged_range in worksheet_values.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        anchor = (min_row, min_col)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_anchor_by_position[(row, col)] = anchor

    header_cells = list(next(worksheet_values.iter_rows(min_row=1, max_row=1), []))
    headers = [to_non_empty_string(cell.value) or "" for cell in header_cells]
    rows: list[dict[str, Any]] = []

    for row_index, row_values in enumerate(
        worksheet_values.iter_rows(min_row=2, max_row=worksheet_values.max_row),
        start=2,
    ):
        row_links = worksheet_links[row_index]
        item: dict[str, Any] = {}
        item["__rowNumber"] = row_index
        for idx, value_cell in enumerate(row_values):
            if idx >= len(headers):
                continue
            header = headers[idx]
            if not header:
                continue
            anchor_position = merged_anchor_by_position.get((row_index, idx + 1))
            if anchor_position and anchor_position != (row_index, idx + 1):
                anchor_row, anchor_col = anchor_position
                resolved_value = worksheet_values.cell(row=anchor_row, column=anchor_col).value
            else:
                resolved_value = value_cell.value

            formula_cell = worksheet_links.cell(
                row=anchor_position[0] if anchor_position else row_index,
                column=anchor_position[1] if anchor_position else (idx + 1),
            )
            formula_text = to_non_empty_string(formula_cell.value)
            if formula_text and formula_text.startswith("="):
                item[f"__formula__{header}"] = formula_text

            if resolved_value is None:
                anchor_row = anchor_position[0] if anchor_position else row_index
                anchor_col = anchor_position[1] if anchor_position else (idx + 1)
                anchor_ref = worksheet_values.cell(row=anchor_row, column=anchor_col).coordinate
                evaluated = _evaluate_text_formula_cell(
                    worksheet_values=worksheet_values,
                    worksheet_links=worksheet_links,
                    cell_ref=anchor_ref,
                    visiting=set(),
                )
                if evaluated is not None:
                    resolved_value = evaluated

            item[header] = resolved_value
            link_cell = row_links[idx] if idx < len(row_links) else None
            if link_cell and link_cell.hyperlink and link_cell.hyperlink.target:
                item[f"__link__{header}"] = str(link_cell.hyperlink.target).strip()
            elif link_cell:
                formula_link, formula_label = _extract_hyperlink_formula_parts(link_cell.value)
                if formula_link:
                    item[f"__link__{header}"] = formula_link
                if item[header] is None and formula_label:
                    item[header] = formula_label
        rows.append(item)

    return rows


def list_offers() -> list[dict[str, Any]]:
    with get_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT
                    id,
                    company,
                    role,
                    applied,
                    archive,
                    status,
                    location,
                    notes,
                    applied_at AS "appliedAt",
                    date_posted AS "datePosted",
                    expires_at AS "expiresAt",
                    CASE
                      WHEN expires_at IS NULL THEN NULL
                      ELSE (expires_at - CURRENT_DATE)
                    END AS "daysToExpire",
                    source,
                    source_url AS "sourceUrl",
                    employment_types AS "employmentTypes",
                    work_time AS "workTime",
                    work_mode AS "workMode",
                    shift_count AS "shiftCount",
                    working_hours AS "workingHours",
                    created_at AS "createdAt"
                FROM applications
                ORDER BY created_at DESC
                LIMIT 500
                """
            )
            rows = cur.fetchall()
            return [dict(row) for row in rows]


def insert_offer(offer: dict[str, Any]) -> dict[str, Any]:
    with get_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                INSERT INTO applications (
                    company, role, applied, archive, status, location, notes, applied_at, date_posted, expires_at, source, source_url,
                    employment_types, work_time, work_mode, shift_count, working_hours
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id, company, role, applied, archive, status, location, notes,
                          applied_at AS "appliedAt", date_posted AS "datePosted", expires_at AS "expiresAt",
                          CASE
                            WHEN expires_at IS NULL THEN NULL
                            ELSE (expires_at - CURRENT_DATE)
                          END AS "daysToExpire",
                          source, source_url AS "sourceUrl",
                          employment_types AS "employmentTypes", work_time AS "workTime",
                          work_mode AS "workMode", shift_count AS "shiftCount",
                          working_hours AS "workingHours", created_at AS "createdAt"
                """,
                (
                    offer["company"],
                    offer["role"],
                    offer.get("applied", True),
                    offer.get("archive", False),
                    normalize_offer_status(offer.get("status"), offer.get("applied") is not False),
                    offer.get("location") or None,
                    offer.get("notes") or None,
                    safe_date(offer.get("appliedAt")),
                    safe_date(offer.get("datePosted")),
                    safe_date(offer.get("expiresAt")),
                    offer.get("source") or None,
                    offer.get("sourceUrl") or None,
                    offer.get("employmentTypes") or None,
                    offer.get("workTime") or None,
                    offer.get("workMode") or None,
                    offer.get("shiftCount") or None,
                    offer.get("workingHours") or None,
                ),
            )
            row = cur.fetchone()
        conn.commit()
        return dict(row) if row else {}


def update_offer(offer_id: int, offer: dict[str, Any]) -> dict[str, Any] | None:
    with get_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                UPDATE applications
                SET
                    company = %s,
                    role = %s,
                    applied = %s,
                    archive = %s,
                    status = %s,
                    location = %s,
                    notes = %s,
                    applied_at = %s,
                    date_posted = %s,
                    expires_at = %s,
                    source = %s,
                    source_url = %s,
                    employment_types = %s,
                    work_time = %s,
                    work_mode = %s,
                    shift_count = %s,
                    working_hours = %s,
                    updated_at = NOW()
                WHERE id = %s
                RETURNING id, company, role, applied, archive, status, location, notes,
                          applied_at AS "appliedAt", date_posted AS "datePosted", expires_at AS "expiresAt",
                          CASE
                            WHEN expires_at IS NULL THEN NULL
                            ELSE (expires_at - CURRENT_DATE)
                          END AS "daysToExpire",
                          source, source_url AS "sourceUrl",
                          employment_types AS "employmentTypes", work_time AS "workTime",
                          work_mode AS "workMode", shift_count AS "shiftCount",
                          working_hours AS "workingHours", created_at AS "createdAt"
                """,
                (
                    offer["company"],
                    offer["role"],
                    offer.get("applied", True),
                    offer.get("archive", False),
                    normalize_offer_status(offer.get("status"), offer.get("applied") is not False),
                    offer.get("location") or None,
                    offer.get("notes") or None,
                    safe_date(offer.get("appliedAt")),
                    safe_date(offer.get("datePosted")),
                    safe_date(offer.get("expiresAt")),
                    offer.get("source") or None,
                    offer.get("sourceUrl") or None,
                    offer.get("employmentTypes") or None,
                    offer.get("workTime") or None,
                    offer.get("workMode") or None,
                    offer.get("shiftCount") or None,
                    offer.get("workingHours") or None,
                    offer_id,
                ),
            )
            row = cur.fetchone()
        conn.commit()
        return dict(row) if row else None


def delete_offer(offer_id: int) -> bool:
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM applications WHERE id = %s", (offer_id,))
            deleted = cur.rowcount > 0
        conn.commit()
    return deleted


def import_offers_from_excel(content: bytes) -> dict[str, Any]:
    preview = preview_offers_from_excel(content)
    mapped_offers = preview.get("offers") or []

    saved: list[dict[str, Any]] = []
    for mapped_offer in mapped_offers:
        saved.append(insert_offer(mapped_offer))

    return {
        "ok": True,
        "imported": len(saved),
        "skipped": preview.get("skipped", 0),
        "ignored": preview.get("ignored", 0),
        "offers": saved,
        "issues": preview.get("issues", []),
    }


def preview_offers_from_excel(content: bytes) -> dict[str, Any]:
    rows = read_excel_rows_with_hyperlinks(content)

    mapped_offers: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    skipped = 0
    ignored = 0
    for row in rows:
        mapped_offer, issue = map_excel_row_to_offer(row, import_source="import_excel")
        if mapped_offer:
            mapped_offers.append(mapped_offer)
        else:
            if issue:
                skipped += 1
                issues.append(issue)
            else:
                ignored += 1

    return {
        "ok": True,
        "imported": len(mapped_offers),
        "skipped": skipped,
        "ignored": ignored,
        "offers": mapped_offers,
        "issues": issues,
    }


def export_offers_to_excel_bytes(user_utc_offset_minutes: int | None = None) -> tuple[bytes, str]:
    offers = list_offers()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "applications"

    headers = [
        "company",
        "role",
        "applied",
        "archive",
        "status",
        "location",
        "notes",
        "appliedAt",
        "datePosted",
        "expiresAt",
        "daysToExpire",
        "source",
        "employmentTypes",
        "workTime",
        "workMode",
        "shiftCount",
        "workingHours",
        "createdAt",
    ]
    worksheet.append(headers)

    for offer in offers:
        raw_date_posted = safe_date(offer.get("datePosted"))
        raw_expires_at = safe_date(offer.get("expiresAt"))
        export_date_posted = raw_date_posted
        # Infer only for export preview/file generation when only closing date exists.
        source_value = str(offer.get("source") or "").strip().lower()
        is_manual_source = source_value in {"", "manual"}
        if not export_date_posted and raw_expires_at and is_manual_source:
            export_date_posted = infer_date_posted_from_expires_at(raw_expires_at)

        worksheet.append(
            [
                offer.get("company") or "",
                offer.get("role") or "",
                offer.get("applied") if offer.get("applied") is not None else True,
                offer.get("archive") is True,
                offer.get("status") or "",
                offer.get("location") or "",
                offer.get("notes") or "",
                safe_date(offer.get("appliedAt")) or "",
                export_date_posted or "",
                raw_expires_at or "",
                offer.get("daysToExpire") if offer.get("daysToExpire") is not None else "",
                offer.get("source") or "",
                ", ".join(offer.get("employmentTypes") or []),
                offer.get("workTime") or "",
                offer.get("workMode") or "",
                offer.get("shiftCount") or "",
                offer.get("workingHours") or "",
                _format_created_at_for_export(offer.get("createdAt"), user_utc_offset_minutes),
            ]
        )
        role_cell = worksheet.cell(row=worksheet.max_row, column=2)
        role_link = to_non_empty_string(offer.get("sourceUrl"))
        if role_link and is_absolute_http_url(role_link):
            role_cell.hyperlink = role_link
            role_cell.style = "Hyperlink"

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"applymanager-offers-{datetime.now(timezone.utc).date().isoformat()}.xlsx"
    return buffer.read(), filename


def get_offer_stats() -> dict[str, Any]:
    offers = list_offers()
    if not offers:
        return {
            "totalOffers": 0,
            "appliedOffers": 0,
            "activeOffers": 0,
            "expiredOffers": 0,
            "averageDaysLeft": None,
            "recentApplications7d": 0,
            "statusCounts": {},
            "sourceCounts": {},
        }

    df = pd.DataFrame(offers)
    total = int(len(df.index))

    applied = int(df["applied"].eq(True).sum()) if "applied" in df.columns else 0

    days_series = pd.to_numeric(df["daysToExpire"], errors="coerce") if "daysToExpire" in df.columns else pd.Series(dtype="float64")
    active = int((days_series >= 0).sum()) if not days_series.empty else 0
    expired = int((days_series < 0).sum()) if not days_series.empty else 0

    status_series = (
        df["status"].fillna("unknown").astype(str).str.strip().replace("", "unknown")
        if "status" in df.columns
        else pd.Series(dtype="string")
    )
    status_counts = {str(key): int(value) for key, value in status_series.value_counts().to_dict().items()}

    source_series = (
        df["source"].fillna("manual").astype(str).str.strip().replace("", "manual")
        if "source" in df.columns
        else pd.Series(dtype="string")
    )
    source_counts = {str(key): int(value) for key, value in source_series.value_counts().to_dict().items()}

    today = pd.Timestamp(datetime.now(timezone.utc).date())
    applied_dates = pd.to_datetime(df["appliedAt"], errors="coerce") if "appliedAt" in df.columns else pd.Series(dtype="datetime64[ns]")
    delta_days = (today - applied_dates.dt.normalize()).dt.days if not applied_dates.empty else pd.Series(dtype="float64")
    recent_applications = int(((delta_days >= 0) & (delta_days <= 7)).sum()) if not delta_days.empty else 0

    avg_days_left = round(float(days_series.mean()), 1) if not days_series.dropna().empty else None

    return {
        "totalOffers": total,
        "appliedOffers": applied,
        "activeOffers": active,
        "expiredOffers": expired,
        "averageDaysLeft": avg_days_left,
        "recentApplications7d": recent_applications,
        "statusCounts": status_counts,
        "sourceCounts": source_counts,
    }
